from pathlib import Path
from openpyxl import load_workbook
import json
import csv
import time
from openai import OpenAI
from typing import List
from dotenv import load_dotenv

OUTPUT_CSV = "tags_classification_by_llm.csv"
DATA_DIR = Path("data/boq_files")        # Excel files
BATCH_SIZE = 30

ALLOWED_TRADES = [
    "Groundworks",
    "Concrete + Formwork",
    "Reinforcement",
    "Structural Steel",
    "Masonry",
    "Waterproofing",
    "Roofing",
    "Mechanical",
    "Plumbing",
    "Electrical",
    "Fire Fighting",
    "HVAC",
    "Drywall / Partitions",
    "Ceilings",
    "Flooring",
    "Tiling",
    "Painting",
    "Joinery",
    "Glazing",
    "Facade",
    "Landscaping",
    "External Works",
    "General / Preliminaries",
]

PROMPT_CONFIG = {
    "system": "You are a construction BOQ classification assistant.",
    "user_template": (
        "Classify the following BOQ TAG into a single allowed trade: {{description}}. "
        "Allowed trades are: {{allowed_trades}}. "
        "Return only one trade or 'UNMAPPED' if none applies."
    )
}

# Load .env from project root
load_dotenv(Path(__file__).parent / ".env")
client = OpenAI()

# ---------------- FUNCTIONS ----------------
def classify_trade(description: str, allowed_trades: List[str], model: str = "gpt-4o-mini") -> str:
    user_prompt = PROMPT_CONFIG["user_template"] \
        .replace("{{description}}", description.strip()) \
        .replace("{{allowed_trades}}", ", ".join(allowed_trades))
    
    response = client.responses.create(
        model=model,
        instructions=PROMPT_CONFIG["system"],
        input=user_prompt
    )
    return response.output_text.strip()

def extract_unique_tags(data_dir: Path) -> List[str]:
    unique_tags = set()
    for excel_path in data_dir.glob("*.xlsx"):
        wb = load_workbook(excel_path, data_only=True)
        for sheet in wb.worksheets:
            tag_col = None
            for row in sheet.iter_rows(min_row=1, max_row=10):
                for cell in row:
                    if cell.value and str(cell.value).strip().upper() == "TAG":
                        tag_col = cell.column
                        break
                if tag_col:
                    break
            if not tag_col:
                continue
            for row in sheet.iter_rows(min_row=2):
                tag = row[tag_col - 1].value
                if tag:
                    unique_tags.add(str(tag).strip())
    return sorted(unique_tags)

# ---------------- MAIN SCRIPT ----------------
def main():
    # Step 1: Extract unique tags
    unique_tags = extract_unique_tags(DATA_DIR)
    print(f"Extracted {len(unique_tags)} unique tags.")

    # Step 2: Classify tags with LLM
    tag_classification = {}
    for i in range(0, len(unique_tags), BATCH_SIZE):
        batch = unique_tags[i:i+BATCH_SIZE]
        print(f"Classifying batch {i//BATCH_SIZE + 1} / {(len(unique_tags)+BATCH_SIZE-1)//BATCH_SIZE}")
        for tag in batch:
            try:
                trade = classify_trade(tag, ALLOWED_TRADES)
                tag_classification[tag] = trade
                # print(f"{tag} -> {trade}")
                time.sleep(0.5)  # avoid rate limits
            except Exception as e:
                print(f"Error classifying {tag}: {e}")
                tag_classification[tag] = "UNMAPPED"

    # Step 3: Save LLM classifications CSV
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["raw_tag", "suggested_allowed_trade"])
        writer.writeheader()
        for tag, trade in tag_classification.items():
            writer.writerow({"raw_tag": tag, "suggested_allowed_trade": trade})

    print(f"LLM classification CSV saved to {OUTPUT_CSV}")

if __name__ == "__main__":
    main()