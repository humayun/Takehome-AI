#!/usr/bin/env python3
import json
import csv
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import load_workbook


def load_boq_items(file_path: str) -> List[Dict[str, Any]]:
    """
    Extract BOQ items from Excel consistently.
    Rules:
    1. Rows with QTY are treated as actual work items.
    2. Preceding rows without QTY are context.
    3. Uppercase-only lines are treated as sticky headings for the next QTY item.
    """
    wb = load_workbook(file_path, data_only=True)
    items: List[Dict[str, Any]] = []

    for sheet in wb.worksheets:
        context_lines: List[str] = []
        sticky_heading: str = None

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            code, description, qty, unit, rate, total, *rest = row + (None,) * (7 - len(row))
            description = str(description).strip() if description else ""

            if not description:
                continue

            if qty is None and description.isupper() and len(description) > 3:
                sticky_heading = description
                continue

            if qty is None:
                context_lines.append(description)
                continue

            full_description = []
            if sticky_heading:
                full_description.append(sticky_heading)
            full_description += context_lines
            full_description.append(description)

            item = {
                "file": Path(file_path).name,
                "sheet": sheet.title,
                "row": row_idx,
                "code": code,
                "description": " | ".join(full_description),
                "qty": qty,
                "unit": unit
            }
            items.append(item)

            context_lines = []

    return items


def save_predictions(predictions: List[Dict[str, Any]], output_path: str = "predictions.json"):
    """Save predictions to JSON file in a readable format."""
    with open(output_path, "w") as f:
        json.dump(predictions, f, indent=2)
    print(f"Saved {len(predictions)} predictions to {output_path}")


def load_allowed_tag(tags_csv):
    # ---------------- LOAD Allowed TAGs ----------------
    allowed_trades = {}
    unique_trades = []
    with open(tags_csv, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            raw_tag = row["raw_tag"].strip()
            allowed_trade = row["suggested_allowed_trade"].strip()
            allowed_trades[raw_tag] = allowed_trade

            if allowed_trade not in unique_trades:
                unique_trades.append(allowed_trade)
    return allowed_trades, unique_trades


def normalize_tag(tag: str) -> str:
    return " ".join(tag.strip().split())


def compare_predictions(predicted_rows, tag_mapping):
    correct = 0
    total = 0
    mismatches = []

    for row in predicted_rows:
        predicted_tag = row.get("predicted_tag", "UNMAPPED").strip()
        item_text = row.get("item", "").strip()

        # Map raw tag from Excel to allowed trade
        mapped_trade = "UNMAPPED"
        for raw_tag, trade in tag_mapping.items():
            if raw_tag.upper() in item_text.upper():
                mapped_trade = trade
                break

        if mapped_trade == "UNMAPPED":
            continue  # skip unmapped tags

        total += 1
        if predicted_tag == mapped_trade:
            correct += 1
        else:
            mismatches.append({
                "item": item_text,
                "predicted_tag": predicted_tag,
                "mapped_trade": mapped_trade
            })

    return total, correct, mismatches